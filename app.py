import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Alignment
import io
import random
import json
import os
from ortools.sat.python import cp_model

# --- 설정 및 전역 변수 ---
SHIFTS = ['D', 'E', 'N', 'O'] # O는 Off(휴무)를 의미 (X, SX, H, HX, TR 등 통합)
SHIFT_IDX = {'D': 0, 'E': 1, 'N': 2, 'O': 3}
REV_SHIFT_IDX = {0: 'D', 1: 'E', 2: 'N', 3: 'X'}

MEN = ["최충일", "윤진호", "이용재"]
LEADER = ["용하영", "최충일", "박세은", "김소은", "윤지선", "이소희", "정하림", "최아라"]
SUB_LEADER = ["김민지", "박우영", "오지은"]

HX_SETTINGS_FILE = "hx_settings.json"

def load_hx_settings():
    """저장된 HX 설정 불러오기"""
    if os.path.exists(HX_SETTINGS_FILE):
        try:
            with open(HX_SETTINGS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            return {}
    return {}

def save_hx_settings(hx_1_5, hx_6_10, hx_11_15, hx_16_20, hx_21_25, hx_26_end):
    """현재 HX 설정 저장하기"""
    with open(HX_SETTINGS_FILE, "w", encoding="utf-8") as f:
        json.dump({
            "hx_1_5": hx_1_5, 
            "hx_6_10": hx_6_10, 
            "hx_11_15": hx_11_15,
            "hx_16_20": hx_16_20,
            "hx_21_25": hx_21_25,
            "hx_26_end": hx_26_end
        }, f, ensure_ascii=False)

def is_cell_colored(cell):
    """셀에 배경색(노란색 등)이 칠해져 있는지 확인"""
    if cell.fill and cell.fill.patternType == 'solid':
        if cell.fill.fgColor.rgb not in ['00000000', 'FFFFFFFF', None]:
            return True
    return False

def parse_shift_options(val_str):
    """쉼표로 구분된 근무 옵션을 인덱스 리스트로 변환"""
    if not val_str or str(val_str).strip().upper() == "NONE":
        return []
    options = []
    for part in str(val_str).split(','):
        p = part.strip().upper()
        if 'D' in p and 'DE' not in p:
            options.append(0)
        elif 'E' in p:
            options.append(1)
        elif 'N' in p:
            options.append(2)
        elif 'X' in p or 'H' in p or 'TR' in p or 'SX' in p or 'O' in p or '기타' in p or '교육' in p:
            options.append(3)
        elif p == 'D':
            options.append(0)
    return list(set(options)) if options else [3]

def get_staff_names(file_content):
    """엑셀 파일에서 직원 이름 목록만 빠르게 추출"""
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb.active
    names = []
    for r in range(3, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if name and str(name).strip() != '이름':
            names.append(str(name).strip())
    return names

def solve_schedule(staff_data, num_days, num_history, base_x_count, min_d, max_d, min_e, max_e, min_n, max_n, custom_rules, weekend_indices):
    model = cp_model.CpModel()
    num_staff = len(staff_data)
    total_days = num_history + num_days
    
    # 변수 생성
    shifts = {}
    is_working = {} 
    for e in range(num_staff):
        for d in range(total_days):
            for s in range(4): # 0:D, 1:E, 2:N, 3:O
                shifts[(e, d, s)] = model.NewBoolVar(f'shift_e{e}_d{d}_s{s}')
            
            # 해당 날짜에 D(0), E(1), N(2) 중 하나라도 하면 1, 아니면 0
            is_working[(e, d)] = model.NewBoolVar(f'work_e{e}_d{d}')
            model.Add(is_working[(e, d)] == sum(shifts[(e, d, s)] for s in [0, 1, 2]))

    objective_terms = []

    # --- 제약 조건 적용 ---
    for e, staff in enumerate(staff_data):
        name = staff['name']

        # [A] 이전달 기록 고정
        for d in range(num_history):
            s_val = staff['history'][d]
            s_idx = SHIFT_IDX.get(s_val, 3)
            model.Add(shifts[(e, d, s_idx)] == 1)

        # [B] 당월 노란색 고정 근무 반영
        for d, opt_list in staff['fixed'].items():
            abs_d = num_history + d
            if opt_list:
                model.Add(sum(shifts[(e, abs_d, s)] for s in opt_list) == 1)

        # [수정사항 1] 원티드/고정에 X가 포함되어 있고 실제로 그 날 X가 배정된다면, 전날 N 절대 불가
        for d in range(num_days):
            abs_d = num_history + d
            opts = staff['fixed'].get(d, [])
            if not opts:
                opts = staff['wanted'].get(d, [])
            
            if 3 in opts: # 원티드에 오프(3)가 포함된 경우
                model.AddImplication(shifts[(e, abs_d, 3)], shifts[(e, abs_d - 1, 2)].Not())

        # [C] 기본 규칙
        for d in range(total_days):
            model.AddExactlyOne(shifts[(e, d, s)] for s in range(4))

        for d in range(total_days - 1):
            # 역방향 교대 금지 (E->D, N->D, N->E 불가)
            model.AddImplication(shifts[(e, d, 1)], shifts[(e, d+1, 0)].Not())
            model.AddImplication(shifts[(e, d, 2)], shifts[(e, d+1, 0)].Not())
            model.AddImplication(shifts[(e, d, 2)], shifts[(e, d+1, 1)].Not())

        for d in range(total_days - 5):
            # 최대 5일 연속 근무 허용
            model.Add(sum(shifts[(e, d+i, 3)] for i in range(6)) >= 1)

        for d in range(total_days - 3):
            # 나이트(N)는 최대 3일 연속까지만
            model.Add(sum(shifts[(e, d+i, 2)] for i in range(4)) <= 3)

        for d in range(total_days - 2):
            # N 근무 후 최소 2일 오프
            model.Add(shifts[(e, d, 2)] + shifts[(e, d+1, 3)] + (1 - shifts[(e, d+2, 3)]) <= 2)

        den_patterns = []
        for d in range(total_days - 2):
            # 3일 연속 D -> E -> N 패턴은 페널티 부여(-40점) (다른 감점들과 합쳐져서 사실상 차단됨)
            den_pattern = model.NewBoolVar(f'den_{e}_{d}')
            model.AddBoolOr([shifts[(e, d, 0)].Not(), shifts[(e, d+1, 1)].Not(), shifts[(e, d+2, 2)].Not(), den_pattern])
            objective_terms.append(-40 * den_pattern) 
            den_patterns.append(den_pattern)
            
        model.Add(sum(den_patterns) <= 2)
        
        for d in range(total_days - 4):
            # 5일 연속 동일 근무(DDDDD, EEEEE) 최대한 자제
            for s in [0, 1]:  
                five_same = model.NewBoolVar(f'five_same_e{e}_d{d}_s{s}')
                model.AddMinEquality(five_same, [shifts[(e, d+i, s)] for i in range(5)])
                objective_terms.append(-80 * five_same)

        for d in range(1, total_days - 1):
            # 독성 퐁당퐁당(Single X) 패턴 절대 금지: E-X-D, N-X-D, N-X-E
            model.AddBoolOr([shifts[(e, d-1, 1)].Not(), shifts[(e, d, 3)].Not(), shifts[(e, d+1, 0)].Not()])
            model.AddBoolOr([shifts[(e, d-1, 2)].Not(), shifts[(e, d, 3)].Not(), shifts[(e, d+1, 0)].Not()])
            model.AddBoolOr([shifts[(e, d-1, 2)].Not(), shifts[(e, d, 3)].Not(), shifts[(e, d+1, 1)].Not()])
            
            # 기타 Single X (가벼운 퐁당퐁당) 자제 (-50점) -> D-X-N은 허용하지만 살짝 자제하는 정도
            single_x = model.NewBoolVar(f'single_x_{e}_{d}')
            model.Add(single_x >= (1 - shifts[(e, d-1, 3)]) + shifts[(e, d, 3)] + (1 - shifts[(e, d+1, 3)]) - 2)
            objective_terms.append(-50 * single_x) 
            
            # 비근무 사이 근무 하나만 껴있는 퐁당퐁당 근무(O-W-O) 최대한 자제 (-80점)
            single_work = model.NewBoolVar(f'single_work_{e}_{d}')
            model.Add(single_work >= shifts[(e, d-1, 3)] + (1 - shifts[(e, d, 3)]) + shifts[(e, d+1, 3)] - 2)
            objective_terms.append(-80 * single_work)
            
            # [핵심 튜닝] D나 E가 하나만 있는 경우 (D-E-E-E, D-D-E 등) -> 약한 감점(-10점)으로 융통성 부여
            for s in [0, 1]: 
                single_shift = model.NewBoolVar(f'single_shift_e{e}_d{d}_s{s}')
                model.Add(single_shift >= shifts[(e, d, s)] + (1 - shifts[(e, d-1, s)]) + (1 - shifts[(e, d+1, s)]) - 2)
                objective_terms.append(-10 * single_shift)

            # [핵심 튜닝] 인터벌 마지막에 N 하나로 끝나는 극악 패턴(D-D-N-O, E-E-N-O) -> 초강력 감점(-100점)
            single_n_end = model.NewBoolVar(f'single_n_end_e{e}_d{d}')
            worked_d_e_before = model.NewBoolVar(f'worked_d_e_before_e{e}_d{d}')
            # 전날에 D 또는 E를 했는가?
            model.Add(worked_d_e_before == shifts[(e, d-1, 0)] + shifts[(e, d-1, 1)])
            # (전날 D/E) -> (오늘 N) -> (내일 O)
            model.Add(single_n_end >= shifts[(e, d, 2)] + shifts[(e, d+1, 3)] + worked_d_e_before - 2)
            objective_terms.append(-100 * single_n_end)

        # [D] 당월 오프 카운트 통일
        fixed_h_tr_count = 0
        for raw_val in staff.get('fixed_raw', {}).values():
            opts = [p.strip().upper() for p in str(raw_val).split(',')]
            if any(p in ['H', 'TR'] or '기타' in p or '교육' in p for p in opts):
                fixed_h_tr_count += 1
                
        target_offs = int(base_x_count) + fixed_h_tr_count
        if not staff['is_male']:
            target_offs += 1

        off_count = sum(shifts[(e, num_history + d, 3)] for d in range(num_days))
        model.Add(off_count == target_offs) # 모든 사람 X 개수 완벽 통일

        # D와 E 갯수 비슷하게 맞추기
        d_count = sum(shifts[(e, num_history + d, 0)] for d in range(num_days))
        e_count = sum(shifts[(e, num_history + d, 1)] for d in range(num_days))
        de_diff = model.NewIntVar(0, num_days, f'de_diff_{e}')
        model.Add(de_diff >= d_count - e_count)
        model.Add(de_diff >= e_count - d_count)
        objective_terms.append(-5 * de_diff)

        # [특이사항 맞춤조건]
        if name in custom_rules.get('weekend_off', []):
            for d in weekend_indices:
                objective_terms.append(-60 * is_working[(e, num_history + d)])
                
        if name in custom_rules.get('weekend_work', []):
            for d in weekend_indices:
                objective_terms.append(60 * is_working[(e, num_history + d)])
                
        if name in custom_rules.get('five_days', []):
            for d in range(total_days - 4):
                five_work = model.NewBoolVar(f'five_work_e{e}_d{d}')
                model.AddMinEquality(five_work, [is_working[(e, d+i)] for i in range(5)])
                objective_terms.append(60 * five_work)
                
        if name in custom_rules.get('no_night', []):
            for d in range(num_days):
                model.Add(shifts[(e, num_history + d, 2)] == 0)


    # [E] 전역 제약 조건 (리더 및 인원수)
    leader_and_sub_indices = [i for i, s in enumerate(staff_data) if s['is_leader'] or s['is_sub_leader']]
    leader_indices = [i for i, s in enumerate(staff_data) if s['is_leader']]
    
    for d in range(num_days):
        abs_d = num_history + d
        for s in [0, 1, 2]: 
            if leader_and_sub_indices:
                model.Add(sum(shifts[(e, abs_d, s)] for e in leader_and_sub_indices) >= 1)
            
            if leader_indices:
                has_leader = model.NewBoolVar(f'has_leader_d{abs_d}_s{s}')
                model.Add(sum(shifts[(e, abs_d, s)] for e in leader_indices) >= 1).OnlyEnforceIf(has_leader)
                model.Add(sum(shifts[(e, abs_d, s)] for e in leader_indices) == 0).OnlyEnforceIf(has_leader.Not())
                objective_terms.append(100 * has_leader)

        d_count_global = sum(shifts[(e, abs_d, 0)] for e in range(num_staff))
        e_count_global = sum(shifts[(e, abs_d, 1)] for e in range(num_staff))
        n_count_global = sum(shifts[(e, abs_d, 2)] for e in range(num_staff))

        model.Add(d_count_global >= int(min_d))
        model.Add(d_count_global <= int(max_d))
        model.Add(e_count_global >= int(min_e))
        model.Add(e_count_global <= int(max_e))
        model.Add(n_count_global >= int(min_n))
        model.Add(n_count_global <= int(max_n))

        # [수정사항 2] 하루 당 E >= N > D 조건 (강제 규칙)
        model.Add(e_count_global >= n_count_global)
        model.Add(n_count_global >= d_count_global + 1)

    # [수정사항 3] 근무 겹침 방지 조합 (A와 B는 같은 날 D, E, N 불가)
    for group in custom_rules.get('not_together', []):
        if len(group) >= 2:
            group_indices = [i for i, s in enumerate(staff_data) if s['name'] in group]
            for d in range(num_days):
                abs_d = num_history + d
                for s in [0, 1, 2]: # D, E, N에 대해서만
                    model.Add(sum(shifts[(e, abs_d, s)] for e in group_indices) <= 1)

    # N(나이트) 개수 균등 분배 
    n_balance_counts = []
    for e, staff in enumerate(staff_data):
        n_count = model.NewIntVar(0, num_days, f'n_count_{e}')
        model.Add(n_count == sum(shifts[(e, num_history + d, 2)] for d in range(num_days)))
        
        if staff['name'] not in custom_rules.get('no_night', []):
            n_balance_counts.append(n_count)
            
    if n_balance_counts:
        max_n_count = model.NewIntVar(0, num_days, 'max_n_count')
        min_n_count = model.NewIntVar(0, num_days, 'min_n_count')
        model.AddMaxEquality(max_n_count, n_balance_counts)
        model.AddMinEquality(min_n_count, n_balance_counts)
        model.Add(max_n_count - min_n_count <= 1)

    # [F] 원티드 반영 (가산점)
    for e, staff in enumerate(staff_data):
        for d, opt_list in staff['wanted'].items():
            abs_d = num_history + d
            for s in opt_list:
                objective_terms.append(15 * shifts[(e, abs_d, s)])
    
    model.Maximize(sum(objective_terms))
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 120.0
    status = solver.Solve(model)

    if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
        for e, staff in enumerate(staff_data):
            for d in range(num_days):
                for s in range(4):
                    if solver.Value(shifts[(e, num_history + d, s)]) == 1:
                        # 무조건 인덱스 번호(s)를 저장하여 'E' 에러를 완벽 차단
                        staff['final_schedule'][d] = s
        return True
    else:
        return False

def process_excel(file_content, target_x_count, min_d, max_d, min_e, max_e, min_n, max_n, custom_rules):
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb.active

    days_of_week = ['월', '화', '수', '목', '금', '토', '일']
    cal_cols = []
    summary_cols = {}
    weekend_indices = []
    
    for col in range(2, ws.max_column + 1):
        v1 = str(ws.cell(row=1, column=col).value).strip().upper() if ws.cell(row=1, column=col).value is not None else ""
        v2 = str(ws.cell(row=2, column=col).value).strip() if ws.cell(row=2, column=col).value is not None else ""
        
        if v2 in days_of_week:
            cal_cols.append(col)
        elif v1 in ['D', 'E', 'N', 'X', 'H', 'HX', 'SX', 'I', '총합']:
            summary_cols[v1] = col

    history_cols = cal_cols[:5] if len(cal_cols) >= 5 else cal_cols[:1]
    current_cols = cal_cols[5:] if len(cal_cols) >= 5 else cal_cols[1:]
    num_history = len(history_cols)
    num_days = len(current_cols)
    
    for d, col in enumerate(current_cols):
        day_str = str(ws.cell(row=2, column=col).value).strip()
        if day_str in ['토', '일']:
            weekend_indices.append(d)

    staff_data = []
    for r in range(3, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if not name or str(name).strip() == '이름':
            continue
            
        name = str(name).strip()
        staff_dict = {
            'row': r,
            'name': name,
            'is_male': name in MEN,
            'is_leader': name in LEADER,
            'is_sub_leader': name in SUB_LEADER,
            'history': [],
            'fixed': {},
            'fixed_raw': {},
            'wanted': {},
            'final_schedule': {}
        }
        
        for col in history_cols:
            val = str(ws.cell(row=r, column=col).value).strip().upper()
            if val in ['D', 'E', 'N']:
                staff_dict['history'].append(val)
            else:
                staff_dict['history'].append('O')
                
        for d, col in enumerate(current_cols):
            cell = ws.cell(row=r, column=col)
            val = str(cell.value).strip()
            
            if val and val.upper() != "NONE":
                opts = parse_shift_options(val)
                if opts:
                    if is_cell_colored(cell):
                        staff_dict['fixed'][d] = opts
                        staff_dict['fixed_raw'][d] = val
                    else:
                        staff_dict['wanted'][d] = opts
                    
        staff_data.append(staff_dict)

    success = solve_schedule(staff_data, num_days, num_history, target_x_count, min_d, max_d, min_e, max_e, min_n, max_n, custom_rules, weekend_indices)

    if not success:
        return None

    blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')
    
    for staff in staff_data:
        r = staff['row']
        counts = {'D':0, 'E':0, 'N':0, 'X':0, 'H':0, 'HX':0, 'SX':0, 'I':0}
        name = staff['name']
        
        # --- HX 시기 고정 로직 (지정된 범위에서 가장 먼저 나오는 오프를 찾음) ---
        hx_target_d = -1
        has_hx_fixed = any('HX' in str(v).upper() for v in staff.get('fixed_raw', {}).values())
        if not staff['is_male'] and not has_hx_fixed:
            all_x_days = [d for d in range(num_days) if staff['final_schedule'].get(d, 3) == 3]
            valid_hx_days = []
            for d in all_x_days:
                if d in staff['fixed']:
                    raw_opts = [p.strip().upper() for p in str(staff['fixed_raw'].get(d, '')).split(',')]
                    if any(p in ['X', 'XX', 'O'] for p in raw_opts): 
                        valid_hx_days.append(d)
                else:
                    valid_hx_days.append(d)
            
            preferred = []
            if name in custom_rules.get('hx_1_5', []): preferred = [d for d in valid_hx_days if 0 <= d <= 4]
            elif name in custom_rules.get('hx_6_10', []): preferred = [d for d in valid_hx_days if 5 <= d <= 9]
            elif name in custom_rules.get('hx_11_15', []): preferred = [d for d in valid_hx_days if 10 <= d <= 14]
            elif name in custom_rules.get('hx_16_20', []): preferred = [d for d in valid_hx_days if 15 <= d <= 19]
            elif name in custom_rules.get('hx_21_25', []): preferred = [d for d in valid_hx_days if 20 <= d <= 24]
            elif name in custom_rules.get('hx_26_end', []): preferred = [d for d in valid_hx_days if 25 <= d]
            
            if preferred:
                hx_target_d = preferred[0]
            elif valid_hx_days: # 지정 범위가 없거나 못 찾았을 경우 대비책
                hx_target_d = valid_hx_days[0]
        # ---------------------------------------------
        
        for d in range(num_days):
            col = current_cols[d]
            is_fixed = d in staff['fixed']
            
            cell = ws.cell(row=r, column=col)
            chosen_idx = staff['final_schedule'].get(d, 3)
            
            if is_fixed:
                raw_opts = [p.strip().upper() for p in str(staff['fixed_raw'].get(d, '')).split(',')]
                val = REV_SHIFT_IDX[chosen_idx]
                for p in raw_opts:
                    if chosen_idx == 0 and 'D' in p and 'DE' not in p: val = p; break
                    elif chosen_idx == 1 and 'E' in p: val = p; break
                    elif chosen_idx == 2 and 'N' in p: val = p; break
                    elif chosen_idx == 3 and (p in ['X', 'SX', 'H', 'HX', 'TR', 'O', 'XX'] or '기타' in p or '교육' in p): val = p; break
                
                # 타겟 날짜라면 HX로 덮어쓰기
                if d == hx_target_d:
                    val = 'HX'
            else:
                val = REV_SHIFT_IDX[chosen_idx]
                # 타겟 날짜라면 HX로 덮어쓰기
                if d == hx_target_d:
                    val = 'HX'
                
                cell.fill = blue_fill 
                
            cell.value = val
            cell.alignment = center_align
                
            raw_cell_val = str(val).strip().upper()
            if raw_cell_val == 'D': counts['D'] = counts.get('D', 0) + 1
            elif raw_cell_val == 'E': counts['E'] = counts.get('E', 0) + 1
            elif raw_cell_val == 'N': counts['N'] = counts.get('N', 0) + 1
            elif raw_cell_val == 'H': counts['H'] = counts.get('H', 0) + 1
            elif raw_cell_val == 'HX': counts['HX'] = counts.get('HX', 0) + 1
            elif raw_cell_val == 'SX': counts['SX'] = counts.get('SX', 0) + 1
            elif 'TR' in raw_cell_val or '기타' in raw_cell_val or '교육' in raw_cell_val: 
                counts['I'] = counts.get('I', 0) + 1 
            elif raw_cell_val in ['X', 'O', 'XX']: 
                counts['X'] = counts.get('X', 0) + 1 
            else: 
                counts['X'] = counts.get('X', 0) + 1
                
        for key, col in summary_cols.items():
            if key == '총합':
                ws.cell(row=r, column=col).value = sum(counts.values())
            elif key in counts:
                ws.cell(row=r, column=col).value = counts.get(key, 0)
            ws.cell(row=r, column=col).alignment = center_align

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# --- Streamlit UI ---
st.set_page_config(page_title="근무표 자동 생성 프로그램", layout="wide")

st.title("🏥 근무표 자동 생성 프로그램")

with st.sidebar:
    st.header("⚙️ 기본 설정")
    target_x = st.number_input("이번 달 기본 오프(X) 개수", min_value=1, max_value=15, value=8, step=1)
    
    st.header("👥 일별 근무자 인원수")
    st.markdown("*(무조건 **E ≥ N > D** 인원으로 배정되므로, 조건에 모순이 생기면 에러가 납니다)*")
    
    col1, col2 = st.columns(2)
    with col1:
        min_d = st.number_input("Day 최소", min_value=1, max_value=10, value=4, step=1)
        min_e = st.number_input("Evening 최소", min_value=1, max_value=10, value=4, step=1)
        min_n = st.number_input("Night 최소", min_value=1, max_value=10, value=4, step=1)
    with col2:
        max_d = st.number_input("Day 최대", min_value=1, max_value=15, value=6, step=1)
        max_e = st.number_input("Evening 최대", min_value=1, max_value=15, value=6, step=1)
        max_n = st.number_input("Night 최대", min_value=1, max_value=15, value=6, step=1)
    
    st.header("📂 파일 업로드")
    uploaded_file = st.file_uploader("원티드 엑셀 파일 업로드 (.xlsx)", type=["xlsx"])

if uploaded_file is not None:
    st.success("✅ 파일 업로드 완료! 데이터를 분석합니다.")
    
    # 엑셀 파일에서 이름 목록을 바로 뽑아서 선택지로 활용
    staff_names = get_staff_names(uploaded_file.getvalue())
    female_staff_names = [name for name in staff_names if name not in MEN]
    
    st.markdown("---")
    st.subheader("💡 직원별 맞춤 조건(특이사항) 설정")
    st.markdown("원하는 직원을 선택하여 특별한 근무 조건을 부여할 수 있습니다. (선택사항)")
    
    col_a, col_b = st.columns(2)
    with col_a:
        rule_weekend_off = st.multiselect("🏝️ 주말(토,일) 오프 위주로 짤 사람", options=staff_names)
        rule_five_days = st.multiselect("🔥 5일 연속근무 위주로 짤 사람", options=staff_names)
    with col_b:
        rule_no_night = st.multiselect("🌙 나이트(N) 전면 배제 (N 0개)", options=staff_names)
        rule_weekend_work = st.multiselect("💪 주말(토,일) 근무 위주로 짤 사람", options=staff_names)
        
    st.markdown("---")
    st.subheader("🚫 근무 겹침 방지 조합")
    st.markdown("같은 날 같은 근무(D, E, N)에 동시에 들어갈 수 없는 직원을 묶어주세요. (최대 3팀)")
    
    col_n1, col_n2, col_n3 = st.columns(3)
    with col_n1:
        rule_not_1 = st.multiselect("겹침 방지 1팀", options=staff_names)
    with col_n2:
        rule_not_2 = st.multiselect("겹침 방지 2팀", options=staff_names)
    with col_n3:
        rule_not_3 = st.multiselect("겹침 방지 3팀", options=staff_names)

    st.markdown("---")
    st.subheader("🌸 보건휴가(HX) 시기 지정")
    st.markdown("1일-5일, 6일-10일, 11일-15일, 16일-20일, 21일-25일, 26일-말일 중 지정할 수 있습니다.")
    
    # 이전에 저장해둔 HX 설정 불러오기
    hx_settings = load_hx_settings()
    default_hx_1_5 = [n for n in hx_settings.get("hx_1_5", []) if n in female_staff_names]
    default_hx_6_10 = [n for n in hx_settings.get("hx_6_10", []) if n in female_staff_names]
    default_hx_11_15 = [n for n in hx_settings.get("hx_11_15", []) if n in female_staff_names]
    default_hx_16_20 = [n for n in hx_settings.get("hx_16_20", []) if n in female_staff_names]
    default_hx_21_25 = [n for n in hx_settings.get("hx_21_25", []) if n in female_staff_names]
    default_hx_26_end = [n for n in hx_settings.get("hx_26_end", []) if n in female_staff_names]
    
    col_hx1, col_hx2, col_hx3 = st.columns(3)
    with col_hx1:
        rule_hx_1_5 = st.multiselect("1일-5일", options=female_staff_names, default=default_hx_1_5)
    with col_hx2:
        rule_hx_6_10 = st.multiselect("6일-10일", options=female_staff_names, default=default_hx_6_10)
    with col_hx3:
        rule_hx_11_15 = st.multiselect("11일-15일", options=female_staff_names, default=default_hx_11_15)
        
    col_hx4, col_hx5, col_hx6 = st.columns(3)
    with col_hx4:
        rule_hx_16_20 = st.multiselect("16일-20일", options=female_staff_names, default=default_hx_16_20)
    with col_hx5:
        rule_hx_21_25 = st.multiselect("21일-25일", options=female_staff_names, default=default_hx_21_25)
    with col_hx6:
        rule_hx_26_end = st.multiselect("26일-말일", options=female_staff_names, default=default_hx_26_end)
        
    # 저장 버튼
    if st.button("💾 현재 보건휴가(HX) 설정 저장"):
        save_hx_settings(rule_hx_1_5, rule_hx_6_10, rule_hx_11_15, rule_hx_16_20, rule_hx_21_25, rule_hx_26_end)
        st.success("✅ 보건휴가(HX) 지정 목록이 정상적으로 저장되었습니다! (다음번 접속 시 자동으로 불러옵니다)")
        
    custom_rules_dict = {
        'weekend_off': rule_weekend_off,
        'five_days': rule_five_days,
        'no_night': rule_no_night,
        'weekend_work': rule_weekend_work,
        'not_together': [rule_not_1, rule_not_2, rule_not_3],
        'hx_1_5': rule_hx_1_5,
        'hx_6_10': rule_hx_6_10,
        'hx_11_15': rule_hx_11_15,
        'hx_16_20': rule_hx_16_20,
        'hx_21_25': rule_hx_21_25,
        'hx_26_end': rule_hx_26_end
    }
    st.markdown("---")
    
    if st.button("🚀 듀티표 생성 시작"):
        with st.spinner("수천만 가지의 경우의 수를 계산하여 최적의 스케줄을 찾고 있습니다. (최대 120초 소요)"):
            try:
                result_file = process_excel(uploaded_file.getvalue(), target_x, min_d, max_d, min_e, max_e, min_n, max_n, custom_rules_dict)
                
                if result_file:
                    st.balloons()
                    st.success("🎉 모든 조건이 만족되는 완벽한 듀티표가 생성되었습니다!")
                    
                    st.download_button(
                        label="📥 완성된 듀티표 다운로드 (.xlsx)",
                        data=result_file,
                        file_name="완성된_듀티표.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    st.info("💡 새로 입력된 듀티는 **연파랑색** 배경으로 표시됩니다.")
                else:
                    st.error("🚨 제약 조건이 너무 빡빡하여 해답을 찾을 수 없습니다! 설정하신 최소/최대 인원이나 오프 개수를 조금 조절해보세요.")
            except Exception as e:
                st.error(f"오류가 발생했습니다. 엑셀 파일의 양식을 확인해주세요. (에러: {e})")
